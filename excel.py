import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from datetime import datetime

# === 1. Load Excel from full path ===
file_path = r"E:\testing\data.xlsx"  # ✅ Your exact file location
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# === 2. Read ToCities from header row (B and C) ===
to_cities = []
for col in range(2, 4):  # Columns B and C
    city = sheet.cell(row=1, column=col).value
    if city:
        to_cities.append((col, city))

# === 3. Read FromCity + Date + Passengers ===
from_data = []
for row in range(2, sheet.max_row + 1):
    from_city = sheet.cell(row=row, column=1).value
    date_str = sheet.cell(row=row, column=4).value
    passengers = sheet.cell(row=row, column=5).value
    if from_city and date_str and passengers:
        from_data.append((row, from_city, date_str, passengers))

# === 4. Setup WebDriver ===
driver = webdriver.Chrome()
driver.maximize_window()

# === 5. Loop through From-To routes ===
for r_idx, from_city, date_str, passengers in from_data:
    for c_idx, to_city in to_cities:
        try:
            driver.get("https://www.cleartrip.com/")
            time.sleep(4)

            # Close mobile/email pop-up if it appears
            try:
                close_btn = driver.find_element(By.XPATH, '//span[contains(text(),"×")]')
                close_btn.click()
                time.sleep(1)
            except:
                pass

            # Enter From City
            from_input = driver.find_element(By.XPATH, '//input[@placeholder="Where from?"]')
            from_input.clear()
            from_input.send_keys(from_city)
            time.sleep(1)
            from_input.send_keys(Keys.DOWN, Keys.RETURN)

            # Enter To City
            to_input = driver.find_element(By.XPATH, '//input[@placeholder="Where to?"]')
            to_input.clear()
            to_input.send_keys(to_city)
            time.sleep(1)
            to_input.send_keys(Keys.DOWN, Keys.RETURN)

            # Click date picker
            date_picker = driver.find_element(By.XPATH, '//div[@data-testid="departure-date-input-container"]')
            date_picker.click()
            time.sleep(1)

            # Convert date format: DD/MM/YYYY → 'Sat, 28 Jun 2025'
            date_obj = datetime.strptime(date_str, "%d/%m/%Y")
            formatted_date = date_obj.strftime("%a, %d %b %Y")

            # Select the correct date
            date_button = driver.find_element(By.XPATH, f'//div[@aria-label="{formatted_date}"]')
            date_button.click()
            time.sleep(1)

            # Select number of passengers
            driver.find_element(By.XPATH, '//div[@data-testid="home-page-travellers"]').click()
            time.sleep(1)

            # Reset adult count to 1
            for _ in range(4):
                try:
                    driver.find_element(By.XPATH, '//button[@aria-label="Decrease adults"]').click()
                    time.sleep(0.2)
                except:
                    break

            # Increase adult count as needed
            for _ in range(passengers - 1):
                driver.find_element(By.XPATH, '//button[@aria-label="Increase adults"]').click()
                time.sleep(0.2)

            # Apply passengers
            driver.find_element(By.XPATH, '//button[span[text()="Apply"]]').click()
            time.sleep(1)

            # ✅ Write success result in Excel
            result = f"{from_city} → {to_city} ✅"
            sheet.cell(row=r_idx, column=c_idx).value = result
            print("✅", result)

        except Exception as e:
            # ❌ Write failure result in Excel
            error = f"{from_city} → {to_city} ❌ {str(e).splitlines()[0][:50]}"
            sheet.cell(row=r_idx, column=c_idx).value = error
            print("❌", error)

        time.sleep(2)

# === 6. Save Excel and close browser ===
wb.save(file_path)
driver.quit()
print(f"\n✅ All Done. Results saved to: {file_path}")
