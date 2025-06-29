import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time

driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://example.com/login")

workbook = openpyxl.load_workbook("E:/testing/data.xlsx")
sheet = workbook.active

for row in range(2, sheet.max_row + 1):
    username = sheet.cell(row=row, column=1).value
    password = sheet.cell(row=row, column=2).value

    driver.get("https://example.com/login")
    time.sleep(1)

    driver.find_element(By.ID, "username").clear()
    driver.find_element(By.ID, "username").send_keys(username)
    driver.find_element(By.ID, "password").clear()
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "submit").click()

    time.sleep(2)

driver.quit()
